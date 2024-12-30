from os.path import join
from pathlib import Path

from openpyxl import load_workbook

"""
version:
    2024.12.30

dependencies:
    pip install openpyxl
"""


def excel_protection_unlock(filepath: str):
    """Excel worksheet protection unlock.

    Unlock locked sheets in the Excel file, save it as a new file
    in the location of the file.
    """
    workbook = load_workbook(filepath)

    # For each worksheet.
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        if sheet.protection.sheet:
            # Disable worksheet protection.
            sheet.protection.disable()

    # Save as new file.
    path = Path(filepath)
    workbook.save(join(path.parent, path.stem + "_unlock" + path.suffix))


if __name__ == "__main__":
    myPath = input("Enter the path to the Excel file you want to unlock:\n")
    excel_protection_unlock(myPath)
